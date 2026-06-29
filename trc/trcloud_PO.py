"""
TRCLOUD Unified Extractor v2.1 (Warehouse)
รวม PO / AP / PV / XExpense / Invoice (IV) ในไฟล์เดียว

ค่าเริ่มต้น: ดึงเฉพาะ PO (ใบสั่งซื้อ) — module อื่นปิดไว้ด้วย ENABLE_* flags
"""

import re
import sys
import html
import time
import os
import json
import requests
import pandas as pd
from typing import Optional
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from requests.adapters import HTTPAdapter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from trcloud_auth import get_cookie

# บังคับ console ให้เป็น UTF-8 (กัน UnicodeEncodeError จากภาษาไทย/emoji บน Windows cp1252)
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except Exception:
        pass

# =============================================================================
# CONFIG — แก้ตรงนี้ก่อนรัน
# -----------------------------------------------------------------------------
# หมายเหตุ: ค่าทั้งหมดตั้งตรงนี้ (ตรงกับ trcloud_runner.py)
#          ถ้ารันผ่าน runner ค่าเหล่านี้จะถูก inject ทับให้อัตโนมัติ
# =============================================================================
DATE_FROM = "2026-01-01"
DATE_TO   = "2026-05-31"

# Cookie จาก auto-login (username/password ใน trcloud_auth.py)
COOKIE = get_cookie()

COMPANY_ID = "25"
PASSKEY    = "6a05946b357765415b4c931d2122a8c8"

SLEEP_SECONDS   = 0.15   # หน่วงระหว่างหน้า (pagination)
MAX_WORKERS     = 8      # จำนวน request พร้อมกันตอนดึง detail
INCLUDE_DETAILS = True
OUTPUT_DIR      = r"C:\Users\Lenovo\Music\Python\Full Excel"

# JSON สำหรับ Warehouse API / Web / App (อ่านโดย NestJS)
OUTPUT_JSON_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "warehouse-app", "apps", "api", "data",
)

# =============================================================================
# MODULE FLAGS — เปิด/ปิดแต่ละ module (ไม่ลบ code เดิม แค่ข้ามตอนรัน)
# =============================================================================
ENABLE_PO        = True   # ใบสั่งซื้อ (Purchase Order) — เป้าหมายหลักของ Warehouse
ENABLE_AP        = False  # ใบวางบิล / ค่าใช้จ่าย
ENABLE_PV        = False  # ใบสำคัญจ่าย
ENABLE_XEXPENSE  = False  # รายจ่ายอื่น (XExpense)
ENABLE_IV        = False  # ใบกำกับภาษีรายได้ (Invoice/IV)
ENABLE_DASHBOARD = False  # Dashboard AP↔PV (ต้องเปิด ENABLE_AP + ENABLE_PV)

# =============================================================================
# SHARED UTILITIES
# =============================================================================
BASE_URL = "https://thaidrill.trcloud.co/application"


def clean_html_text(raw_text: str) -> str:
    if not raw_text or not isinstance(raw_text, str):
        return ""
    text = raw_text
    text = re.sub(r'(?i)<script\b[^<]*(?:(?!</script>)<[^<]*)*</script>', '', text)
    text = re.sub(r'(?i)<style\b[^<]*(?:(?!</style>)<[^<]*)*</style>', '', text)
    text = re.sub(r'(?i)<\s*br\s*/?\s*>', '\n', text)
    text = re.sub(r'(?i)</\s*(p|div|li|tr)\s*>', '\n', text)
    text = re.sub(r'<[^>]+>', '', text)
    text = html.unescape(text)
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', text)
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\n\s*\n+', '\n', text)
    return text.strip()


def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "X-Requested-With": "XMLHttpRequest",
        "Cookie": COOKIE,
    })
    # ขยาย connection pool ให้พอกับจำนวน worker ที่ยิงพร้อมกัน
    adapter = HTTPAdapter(pool_connections=MAX_WORKERS, pool_maxsize=MAX_WORKERS)
    s.mount("https://", adapter)
    s.mount("http://", adapter)
    return s


SESSION = make_session()

TEXT_COLS = ['description', 'remark', 'note', 'name', 'organization', 'title']


def post_form(url: str, payload: dict, json_wrap: bool = False, timeout: int = 60) -> dict:
    """POST helper — รองรับทั้ง form-urlencoded และ json= wrap"""
    data = {"json": json.dumps(payload)} if json_wrap else payload
    r = SESSION.post(url, data=data, timeout=timeout)
    r.raise_for_status()
    return r.json()


def base_payload(**overrides) -> dict:
    """payload พื้นฐานที่ใช้ร่วมกันทุก module"""
    p = {
        "company_id": COMPANY_ID, "passkey": PASSKEY,
        "start": 0, "keyword": "", "filter": "",
        "from": DATE_FROM, "to": DATE_TO,
        "date_from": DATE_FROM, "date_to": DATE_TO,
        "activate_date": "on", "department": "", "sort": "",
        "advance_search": "1", "project": "", "staff": "",
        "source": "", "title": "", "name": "", "organization": "",
        "tax_id": "", "doc_from": "", "doc_to": "",
        "total_from": "", "total_to": "", "gtotal_from": "", "gtotal_to": "",
        "vat": "all", "type": "",
    }
    p.update(overrides)
    return p


def paginate_list(url: str, id_key: str, type_param: str = "",
                  json_wrap: bool = False, extra: dict = None) -> list:
    """วน page จนครบ total — ใช้ได้กับ AP/PV"""
    records, seen, page, total = [], set(), 0, None
    while True:
        payload = base_payload(start=page, type=type_param)
        if extra:
            payload.update(extra)
        res = post_form(url, payload, json_wrap=json_wrap)
        if res.get("success") != 1:
            break
        if total is None:
            total = int(res.get("count", 0) or 0)
        items = res.get("result", []) or []
        if not items:
            break
        for it in items:
            pid = it.get(id_key) or it.get("id")
            if pid and pid not in seen:
                seen.add(pid)
                records.append(it)
        if total and len(records) >= total:
            break
        if len(items) < 50:
            break
        page += 1
        time.sleep(SLEEP_SECONDS)
    return records


def clean_list_rows(records: list, priority_cols: list) -> list:
    """clean HTML text + reorder columns"""
    out = []
    for row in records:
        r = row.copy()
        for col in TEXT_COLS:
            if r.get(col) is not None:
                r[col] = clean_html_text(str(r[col]))
        cf  = str(r.get("company_format", "") or "")
        inv = str(r.get("invoice_number", "") or "")
        r["doc_number"] = (cf + inv) if cf else inv
        ordered = {c: r[c] for c in priority_cols if c in r}
        for c in r:
            if c not in ordered:
                ordered[c] = r[c]
        out.append(ordered)
    return out


def fetch_details_parallel(records: list, worker, label: str) -> list:
    """ยิง detail แบบขนานด้วย ThreadPool — worker(rec) คืน list ของ line items
    (และอาจแก้ค่าใน rec ได้โดยตรง เพราะแต่ละ thread แตะ rec ของตัวเองเท่านั้น)."""
    lines, done, total = [], 0, len(records)
    if not records:
        return lines
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futures = [ex.submit(worker, rec) for rec in records]
        for fut in as_completed(futures):
            try:
                lines.extend(fut.result() or [])
            except Exception:
                pass
            done += 1
            if done % 25 == 0 or done == total:
                print(f"   📥 {label} {done}/{total} ({len(lines)} items)")
    return lines


# =============================================================================
# MODULE: PO (ใบสั่งซื้อ / Purchase Order)
# =============================================================================
API_PO_LIST     = f"{BASE_URL}/expense/api/engine-po/po_search_keyword.php"
API_PO_RETRIEVE = f"{BASE_URL}/expense/api/engine-po/retrieve_po.php"

PO_PRIORITY = [
    'doc_number', 'document_number', 'po_id',
    'issue_date', 'due_date', 'reference',
    'name', 'organization', 'tax_id',
    'grand_total', 'before_vat', 'discount', 'vat',
    'status', 'approve_status', 'payment',
    'department', 'project', 'staff', 'source',
    'description', 'remark', 'note',
    'company_format', 'company_id',
]


def format_po_ref(row: dict) -> str:
    """เลขที่ PO รวม prefix เช่น PO26050364"""
    cf = str(row.get("company_format") or "PO").strip()
    num = str(
        row.get("document_number")
        or row.get("invoice_number")
        or row.get("doc_number")
        or ""
    ).strip()
    if not num:
        return cf
    ref = f"{cf}{num}"
    if not ref.upper().startswith("PO"):
        ref = f"PO{num}"
    return ref


def _po_worker(po: dict) -> list:
    """ดึง line items ของ PO หนึ่งใบ (เรียกขนาน)"""
    pid = po.get("po_id") or po.get("id")
    if not pid:
        return []
    lines = []
    try:
        payload = base_payload(type="po")
        payload["id"] = str(pid)
        d = post_form(API_PO_RETRIEVE, payload, json_wrap=True)
        if d.get("success") == 1:
            head = d.get("head") or {}
            po_doc = format_po_ref({**po, **head})
            for line in (d.get("detail") or d.get("item") or d.get("items") or []):
                desc = line.get("description") or ""
                product_name = clean_html_text(str(desc)) if desc else ""
                lines.append({
                    "po_id": pid,
                    "po_ref": po_doc,
                    "document_number": po.get("document_number") or head.get("document_number"),
                    "issue_date": po.get("issue_date") or head.get("issue_date"),
                    "organization": po.get("organization") or head.get("organization"),
                    "product_id": line.get("product_id"),
                    "product_name": product_name,
                    "description": product_name,
                    "unit": line.get("unit") or line.get("sunit"),
                    "quantity": line.get("quantity"),
                    "price": line.get("price"),
                    "discount": line.get("discount"),
                    "before_vat": line.get("before_vat"),
                    "vat": line.get("vat"),
                    "item_total": line.get("total"),
                    "item_id": line.get("item_id"),
                    "po_item_id": line.get("po_item_id") or line.get("iv_item_id"),
                })
    except Exception:
        pass
    return lines


def fetch_po_module():
    print("🚀 [PO] ดึง list...")
    SESSION.headers.update({
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "Referer": f"{BASE_URL}/expense/po.php",
    })
    # PO list ใช้ type=project + activate_date=off (ตามหน้า po.php จริง)
    po_list = paginate_list(
        API_PO_LIST, id_key="po_id", type_param="project",
        json_wrap=False, extra={"activate_date": "off"},
    )
    if po_list and all(not r.get("po_id") for r in po_list):
        for r in po_list:
            r.setdefault("po_id", r.get("expense_id") or r.get("id"))

    po_lines = fetch_details_parallel(po_list, _po_worker, "PO detail") if INCLUDE_DETAILS else []
    print(f"✅ PO: {len(po_list)} ใบ | {len(po_lines)} line items")
    return clean_list_rows(po_list, PO_PRIORITY), po_lines


# =============================================================================
# MODULE: AP (ใบวางบิล/ค่าใช้จ่าย)
# =============================================================================
API_AP_LIST     = f"{BASE_URL}/expense/api/engine-expense/expense_search_keyword.php"
API_AP_RETRIEVE = f"{BASE_URL}/expense/api/engine-expense/retrieve_expense.php"
API_AP_PAYMENT  = f"{BASE_URL}/expense/api/engine-expense/invoice-payment.php"


def _ap_worker(ap: dict) -> list:
    """ดึงประวัติชำระ + line items ของ AP หนึ่งใบ (เรียกขนาน)"""
    eid = ap.get("expense_id") or ap.get("id")
    if not eid:
        return []

    # ประวัติการชำระ
    try:
        res  = post_form(API_AP_PAYMENT, base_payload(activate_date="on", expense_id=eid), json_wrap=True)
        pays = res.get("result", []) if res.get("success") == 1 else []
        if pays:
            ap["payment_doc"]    = ", ".join(
                f"{p.get('company_format','')}{p.get('invoice_number','')}" for p in pays)
            ap["payment_amount"] = sum(float(p.get("payment", 0) or 0) for p in pays)
            ap["payment_status"] = "ชำระแล้ว"
        else:
            ap["payment_status"] = "ยังไม่ชำระ"
            ap["payment_doc"]    = ""
            ap["payment_amount"] = 0.0
    except Exception:
        ap["payment_status"] = "ERR"

    # เลขที่ AP = company_format + invoice_number (เช่น AP26050210)
    ap_doc = f"{ap.get('company_format','')}{ap.get('invoice_number','')}" or ap.get("invoice_ref_no", "")

    # line items — endpoint นี้ต้องส่งแบบ json_wrap + type=ap และ item อยู่ใต้คีย์ 'detail'
    lines = []
    if INCLUDE_DETAILS:
        try:
            payload = base_payload(type="ap")
            payload["id"] = str(eid)
            d = post_form(API_AP_RETRIEVE, payload, json_wrap=True)
            if d.get("success") == 1:
                for line in (d.get("detail") or d.get("item") or d.get("items") or []):
                    lines.append({"ap_doc": ap_doc, "expense_id": eid, **line})
        except Exception:
            pass
    return lines


def fetch_ap_module():
    print("🚀 [AP] ดึง list...")
    ap_list = paginate_list(API_AP_LIST, id_key="expense_id", type_param="ap")
    print(f"   → ได้ {len(ap_list)} ใบ")

    ap_lines = fetch_details_parallel(ap_list, _ap_worker, "AP detail")

    print(f"✅ AP: {len(ap_list)} ใบ | {len(ap_lines)} line items")
    return ap_list, ap_lines


# =============================================================================
# MODULE: PV (ใบสำคัญจ่าย)
# =============================================================================
API_PV_LIST     = f"{BASE_URL}/finance/api/engine-payment/payment_search_keyword.php"
API_PV_RETRIEVE = f"{BASE_URL}/finance/api/engine-payment/retrieve_payment.php"


def _pv_worker(pv: dict) -> list:
    """ดึง detail (PV→AP mapping) ของ PV หนึ่งใบ (เรียกขนาน)"""
    pid = pv.get("payment_id")
    if not pid:
        return []
    lines = []
    try:
        d = post_form(API_PV_RETRIEVE,
                      {"company_id": COMPANY_ID, "passkey": PASSKEY, "id": pid},
                      json_wrap=True)
        details = d.get("detail", []) or []
        pv_doc = f"{pv.get('company_format','')}{pv.get('document_number','')}"
        pv["paid_count"] = len(details)
        pv["paid_docs"]  = ", ".join(
            f"{x.get('format','')}{x.get('document_number','')}" for x in details)
        for line in details:
            lines.append({
                "pv_id": pid,
                "pv_doc": pv_doc,
                "pv_date": pv.get("issue_date"),
                "pv_grand_total": pv.get("grand_total"),
                "organization": pv.get("organization"),
                "paid_docs": f"{line.get('format','')}{line.get('document_number','')}",
                "ap_issue_date": line.get("issue_date", ""),
                "ap_due_date": line.get("due_date", ""),
                "amount": float(line.get("amount", 0) or 0),
                "doc_type": line.get("doc_type", ""),
            })
    except Exception:
        pass
    return lines


def fetch_pv_module():
    print("🚀 [PV] ดึง list...")
    pv_list = paginate_list(API_PV_LIST, id_key="payment_id", json_wrap=True)
    print(f"   → ได้ {len(pv_list)} ใบ")

    pv_lines = fetch_details_parallel(pv_list, _pv_worker, "PV→AP Mapping")

    print(f"✅ PV: {len(pv_list)} ใบ | {len(pv_lines)} mapping rows")
    return pv_list, pv_lines


# =============================================================================
# MODULE: XExpense (รายจ่ายอื่น)
# =============================================================================
API_XEX_LIST     = f"{BASE_URL}/expense/api/engine-expense/expense_search_keyword.php"
API_XEX_RETRIEVE = f"{BASE_URL}/expense/api/engine-expense/retrieve_xexpense.php"

XEX_PRIORITY = [
    'doc_number', 'company_format', 'invoice_number',
    'issue_date', 'due_date', 'reference',
    'name', 'organization', 'tax_id',
    'total', 'grand_total', 'discount', 'tax', 'wht',
    'status', 'payment', 'type', 'expense_type',
    'department', 'project', 'staff', 'source',
    'description', 'remark', 'note',
    'expense_id', 'company_id',
]


def _xex_worker(rec: dict) -> list:
    """ดึง line items ของ XExpense หนึ่งใบ (เรียกขนาน)"""
    eid = rec.get("expense_id")
    if not eid:
        return []
    lines = []
    try:
        payload = base_payload(type="exp")
        payload["id"] = str(eid)
        data = post_form(API_XEX_RETRIEVE, payload, json_wrap=True)
        if data and data.get("success") == 1:
            head    = data.get("head") or {}
            details = data.get("detail") or []
            inv_no  = head.get("invoice_number") or rec.get("invoice_number") or ""
            cf      = head.get("company_format") or rec.get("company_format") or ""
            doc_no  = f"{cf}{inv_no}" if cf else str(inv_no)
            for it in details:
                desc = it.get("description") or ""
                lines.append({
                    'doc_number':    doc_no,
                    'invoice_number': inv_no,
                    'expense_id':    eid,
                    'issue_date':    head.get('issue_date')    or rec.get('issue_date'),
                    'due_date':      head.get('due_date')      or rec.get('due_date'),
                    'name':          head.get('name')          or rec.get('name'),
                    'organization':  head.get('organization')  or rec.get('organization'),
                    'department':    head.get('department')    or rec.get('department'),
                    'project':       head.get('project')       or rec.get('project'),
                    'status':        head.get('status')        or rec.get('status'),
                    'payment':       head.get('payment')       or rec.get('payment'),
                    'acc_code':      it.get('acc_code'),
                    'acc_th':        it.get('acc_th'),
                    'acc_en':        it.get('acc_en'),
                    'description':   clean_html_text(str(desc)) if desc else '',
                    'item_total':    it.get('total'),
                    'x_id':          it.get('x_id'),
                })
    except Exception:
        pass
    return lines


def fetch_xexpense_module():
    print("🚀 [XExpense] ดึง list...")
    records, seen, page, total = [], set(), 0, None

    while True:
        payload = base_payload(start=page, type="exp", vat="all")
        try:
            SESSION.headers.update({
                "Content-Type": "application/x-www-form-urlencoded",
                "Referer": f"{BASE_URL.replace('/application','')}/application/expense/xexpense.php",
            })
            res = post_form(API_XEX_LIST, payload, json_wrap=False)
        except Exception as e:
            print(f"❌ XExpense list error: {e}")
            break

        if res.get("success") != 1:
            print(f"❌ XExpense API: {res.get('message', 'Unknown')}")
            break

        if total is None:
            total = int(res.get("count", 0) or 0)
        items = res.get("result", []) or []
        if not items:
            break
        new_rows = [r for r in items if r.get("expense_id") not in seen]
        for r in new_rows:
            seen.add(r.get("expense_id"))
        records.extend(new_rows)
        print(f"   page {page} | got {len(items)} | new {len(new_rows)} | total {len(records)}/{total}")
        if total and len(records) >= total:
            break
        if len(items) < 50:
            break
        page += 1
        time.sleep(SLEEP_SECONDS)

    # ดึง detail แบบขนาน
    xex_lines = fetch_details_parallel(records, _xex_worker, "XExpense detail") if INCLUDE_DETAILS else []

    print(f"✅ XExpense: {len(records)} ใบ | {len(xex_lines)} line items")
    return clean_list_rows(records, XEX_PRIORITY), xex_lines


# =============================================================================
# MODULE: Invoice / IV (ใบกำกับภาษี รายได้)
# =============================================================================
API_IV_LIST     = f"{BASE_URL}/revenue/api/engine-invoice/invoice_search_keyword.php"
API_IV_RETRIEVE = f"{BASE_URL}/revenue/api/engine-invoice/retrieve_invoice.php"

IV_PRIORITY = [
    'doc_number', 'company_format', 'invoice_number',
    'issue_date', 'due_date', 'reference',
    'name', 'organization', 'tax_id',
    'total', 'grand_total', 'discount', 'tax', 'wht',
    'status', 'payment', 'type',
    'department', 'project', 'staff', 'source',
    'description', 'remark', 'note',
    'invoice_id', 'company_id',
]


def _iv_worker(rec: dict) -> list:
    """ดึง line items ของ Invoice หนึ่งใบ (เรียกขนาน)"""
    iid = rec.get("invoice_id")
    if not iid:
        return []
    lines = []
    try:
        payload = base_payload()
        payload["id"] = str(iid)
        data = post_form(API_IV_RETRIEVE, payload, json_wrap=True)
        if data and data.get("success") == 1:
            head    = data.get("head") or {}
            details = data.get("detail") or []
            inv_no  = head.get("invoice_number") or rec.get("invoice_number") or ""
            cf      = head.get("company_format") or rec.get("company_format") or ""
            doc_no  = f"{cf}{inv_no}" if cf else str(inv_no)
            for it in details:
                desc = it.get("description") or ""
                lines.append({
                    'doc_number':    doc_no,
                    'invoice_number': inv_no,
                    'invoice_id':    iid,
                    'issue_date':    head.get('issue_date')   or rec.get('issue_date'),
                    'due_date':      head.get('due_date')     or rec.get('due_date'),
                    'name':          head.get('name')         or rec.get('name'),
                    'organization':  head.get('organization') or rec.get('organization'),
                    'department':    head.get('department')   or rec.get('department'),
                    'project':       head.get('project')      or rec.get('project'),
                    'status':        head.get('status')       or rec.get('status'),
                    'payment':       head.get('payment')      or rec.get('payment'),
                    'acc_code':      it.get('acc_code'),
                    'acc_th':        it.get('acc_th'),
                    'acc_en':        it.get('acc_en'),
                    'description':   clean_html_text(str(desc)) if desc else '',
                    'item_total':    it.get('total'),
                    'x_id':          it.get('x_id'),
                })
    except Exception:
        pass
    return lines


def fetch_invoice_module():
    print("🚀 [Invoice/IV] ดึง list...")
    records, seen, page, total = [], set(), 0, None

    while True:
        payload = base_payload(start=page)
        payload.pop("type", None)  # IV ไม่ต้องส่ง type
        try:
            SESSION.headers.update({
                "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
                "Referer": f"{BASE_URL.replace('/application','')}/application/revenue/revenue.php",
            })
            res = post_form(API_IV_LIST, payload, json_wrap=False)
        except Exception as e:
            print(f"❌ Invoice list error: {e}")
            break

        if res.get("success") != 1:
            print(f"❌ Invoice API: {res.get('message', 'Unknown')}")
            break

        if total is None:
            total = int(res.get("count", 0) or 0)
            print(f"   → Invoice ทั้งหมด {total} ใบ")
        items = res.get("result", []) or []
        if not items:
            break

        new_rows = [r for r in items if r.get("invoice_id") not in seen]
        for r in new_rows:
            seen.add(r.get("invoice_id"))
        records.extend(new_rows)
        pct = (len(records) / total * 100) if total else 0
        print(f"   page {page} | got {len(items)} | new {len(new_rows)} | {len(records)}/{total} ({pct:.1f}%)")

        if total and len(records) >= total:
            break
        if not new_rows:
            break
        page += 1
        time.sleep(SLEEP_SECONDS)

    # ดึง detail แบบขนาน
    iv_lines = fetch_details_parallel(records, _iv_worker, "Invoice detail") if INCLUDE_DETAILS else []

    print(f"✅ Invoice/IV: {len(records)} ใบ | {len(iv_lines)} line items")
    return clean_list_rows(records, IV_PRIORITY), iv_lines


# =============================================================================
# EXPORT: Raw Data Excel
# =============================================================================
def style_workbook(wb):
    """Apply consistent header styling to all sheets"""
    header_fill  = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font  = Font(color="FFFFFF", bold=True, size=11)
    align_center = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'),
    )
    number_fmt = '#,##0.00'

    for ws in wb.worksheets:
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), 1):
            header_cell = ws.cell(row=1, column=col_idx)
            header_cell.fill      = header_fill
            header_cell.font      = header_font
            header_cell.alignment = align_center
            header_cell.border    = thin_border

            header_val  = str(header_cell.value or "").lower()
            is_numeric  = any(k in header_val for k in ['amount', 'total', 'balance', 'grand'])
            max_length  = len(str(header_cell.value or ""))

            for cell in col_cells[1:]:
                cell.border = thin_border
                if cell.value is not None:
                    try:
                        cl = len(str(cell.value))
                        if cl > max_length:
                            max_length = cl
                    except Exception:
                        pass
                if is_numeric and isinstance(cell.value, (int, float)):
                    cell.number_format = number_fmt

            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 40)

    return wb


def export_master_excel(datasets: dict, timestamp: str,
                        filename_prefix: str = "TRCLOUD_Master_Export") -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, f"{filename_prefix}_{timestamp}.xlsx")
    print(f"\n💾 บันทึก Raw Data → {out_path}")

    priority_cols = [
        "po_id", "document_number", "expense_id", "invoice_id", "payment_id",
        "pv_doc", "pv_date", "organization", "pv_grand_total", "paid_docs",
        "ap_issue_date", "doc_number", "issue_date",
        "grand_total", "payment_status", "status", "approve_status",
    ]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name, data in datasets.items():
            df = pd.DataFrame(data) if data else pd.DataFrame()
            if not df.empty:
                cols = [c for c in priority_cols if c in df.columns] + \
                       [c for c in df.columns if c not in priority_cols]
                df = df[cols]
            # เขียนชีตเสมอ แม้ว่างเปล่า (กันชีตหายเงียบ เช่น Invoice เมื่อ session หมดอายุ)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"   📄 {sheet_name}: {len(df)} rows")

    wb = style_workbook(load_workbook(out_path))
    wb.save(out_path)
    return out_path


def export_po_json(po_list: list, po_lines: list) -> str:
    """บันทึก PO เป็น JSON สำหรับ Warehouse API / Web / App"""
    os.makedirs(OUTPUT_JSON_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_JSON_DIR, "po.json")

    def norm_order(row: dict) -> dict:
        po_ref = format_po_ref(row)
        return {
            "po_id":           str(row.get("po_id") or row.get("id") or ""),
            "po_ref":          po_ref,
            "supplier_name":   str(row.get("organization") or row.get("name") or "").strip(),
            "issue_date":      row.get("issue_date") or "",
            "due_date":        row.get("due_date") or "",
            "grand_total":     float(row.get("grand_total") or row.get("total") or 0),
            "status":          str(row.get("status") or ""),
            "approve_status":  str(row.get("approve_status") or ""),
            "payment":         str(row.get("payment") or ""),
            "department":      str(row.get("department") or ""),
            "project":         str(row.get("project") or ""),
            "reference":       str(row.get("reference") or ""),
            "description":     str(row.get("description") or ""),
            "products":        [],
        }

    orders = [norm_order(r) for r in po_list]
    for o in orders:
        po_lines_matched = [
            ln for ln in po_lines if str(ln.get("po_id", "")) == o["po_id"]
        ]
        o["line_count"] = len(po_lines_matched)
        o["products"] = [
            str(ln.get("product_name") or ln.get("description") or "").strip()
            for ln in po_lines_matched
            if ln.get("product_name") or ln.get("description")
        ]

    payload = {
        "fetched_at": datetime.now().isoformat(timespec="seconds"),
        "date_from":  DATE_FROM,
        "date_to":    DATE_TO,
        "source":     "trcloud",
        "count":      len(orders),
        "orders":     orders,
        "lines":      po_lines,
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, default=str)
    print(f"\n📦 บันทึก JSON → {out_path} ({len(orders)} PO)")
    return out_path


# =============================================================================
# EXPORT: Dashboard Data Model
# =============================================================================
def generate_dashboard_model(input_path: str, output_path: str):
    print(f"\n📊 สร้าง Dashboard Data Model → {output_path}")
    try:
        ap      = pd.read_excel(input_path, sheet_name='AP_Summary')
        pv      = pd.read_excel(input_path, sheet_name='PV_Summary')
        mapping = pd.read_excel(input_path, sheet_name='PV_to_AP_Mapping')
    except Exception as e:
        print(f"⚠️ ไม่สามารถโหลดชีตสำหรับ Dashboard: {e}")
        return

    # doc keys
    ap['AP_Doc'] = 'AP' + ap.get('invoice_number',  pd.Series(dtype=str)).astype(str)
    pv['PV_Doc'] = 'PV' + pv.get('document_number', pd.Series(dtype=str)).astype(str)
    mapping['PV_Doc'] = mapping.get('pv_doc',    pd.Series(dtype=str)).astype(str)
    mapping['AP_Doc'] = mapping.get('paid_docs', pd.Series(dtype=str)).astype(str)

    def safe_cols(df, cols):
        return df[[c for c in cols if c in df.columns]]

    ap_clean  = safe_cols(ap, ['AP_Doc', 'issue_date', 'organization', 'grand_total',
                                'payment_status', 'due_date', 'reference', 'department', 'project'])
    pv_clean  = safe_cols(pv, ['PV_Doc', 'issue_date', 'organization', 'grand_total', 'status'])
    map_clean = safe_cols(mapping, ['pv_doc', 'pv_date', 'organization', 'paid_docs',
                                     'ap_issue_date', 'ap_due_date', 'amount'])

    # AP Debt
    if not map_clean.empty and not ap_clean.empty:
        paid = mapping.groupby('AP_Doc')['amount'].sum().reset_index()
        paid.rename(columns={'amount': 'Total_Paid'}, inplace=True)
        ap_debt = pd.merge(ap_clean, paid, on='AP_Doc', how='left')
        ap_debt['Total_Paid']          = ap_debt['Total_Paid'].fillna(0)
        ap_debt['Outstanding_Balance'] = (ap_debt['grand_total'] - ap_debt['Total_Paid']).round(2)
        ap_debt.sort_values('Outstanding_Balance', ascending=False, inplace=True)
    else:
        ap_debt = pd.DataFrame()

    # Mapping flow: AP → PV
    pv_by_ap = mapping.groupby('AP_Doc')['PV_Doc'].apply(lambda x: list(x.dropna().unique())).to_dict()

    rows = []
    for _, r in ap.iterrows():
        ap_doc = r.get('AP_Doc', '')
        pvs = pv_by_ap.get(ap_doc, [])
        rows.append({'AP_Doc': ap_doc,
                     'Linked_PVs': ', '.join(sorted(set(pvs))),
                     'Document_Date': r.get('issue_date', ''),
                     'Vendor': r.get('organization', ''),
                     'Amount': r.get('grand_total', 0),
                     'Status': r.get('payment_status', '')})

    flow_df = pd.DataFrame(rows)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        flow_df.to_excel(writer,    sheet_name='AP_to_PV_Mapping_Flow', index=False)
        if not ap_debt.empty:
            ap_debt.to_excel(writer, sheet_name='Fact_AP_Debt_Status',  index=False)
        ap_clean.to_excel(writer,   sheet_name='Fact_AP',               index=False)
        pv_clean.to_excel(writer,   sheet_name='Fact_PV',               index=False)
        map_clean.to_excel(writer,  sheet_name='Fact_PV_AP_Mapping',    index=False)

    wb = style_workbook(load_workbook(output_path))
    wb.save(output_path)
    print(f"✅ Dashboard saved.")


# =============================================================================
# MAIN
# =============================================================================
if __name__ == "__main__":
    active_modules = [
        name for name, on in [
            ("PO", ENABLE_PO),
            ("AP", ENABLE_AP),
            ("PV", ENABLE_PV),
            ("XExpense", ENABLE_XEXPENSE),
            ("Invoice/IV", ENABLE_IV),
        ] if on
    ]
    total_steps = len(active_modules)

    print("=" * 60)
    print("  TRCLOUD Unified Extractor v2.1 (Warehouse)")
    print(f"  ช่วง: {DATE_FROM} → {DATE_TO}")
    print(f"  เปิดใช้: {', '.join(active_modules) or '(ไม่มี module ที่เปิด)'}")
    print("=" * 60)

    if not active_modules:
        print("⚠️ ไม่มี module ที่เปิด — ตั้ง ENABLE_PO=True หรือเปิด module อื่นใน CONFIG")
        raise SystemExit(1)

    datasets = {}
    step = 0
    po_list, po_lines = [], []

    if ENABLE_PO:
        step += 1
        print(f"\n[{step}/{total_steps}] PO (ใบสั่งซื้อ)...")
        po_list, po_lines = fetch_po_module()
        datasets["PO_Summary"]   = po_list
        datasets["PO_LineItems"] = po_lines

    if ENABLE_AP:
        step += 1
        print(f"\n[{step}/{total_steps}] AP (ใบวางบิล)...")
        ap_list, ap_lines = fetch_ap_module()
        datasets["AP_Summary"]   = ap_list
        datasets["AP_LineItems"] = ap_lines

    if ENABLE_PV:
        step += 1
        print(f"\n[{step}/{total_steps}] PV (ใบสำคัญจ่าย)...")
        pv_list, pv_lines = fetch_pv_module()
        datasets["PV_Summary"]       = pv_list
        datasets["PV_to_AP_Mapping"] = pv_lines

    if ENABLE_XEXPENSE:
        step += 1
        print(f"\n[{step}/{total_steps}] XExpense (รายจ่ายอื่น)...")
        xex_list, xex_lines = fetch_xexpense_module()
        datasets["XExpense_Summary"] = xex_list
        datasets["XExpense_Items"]   = xex_lines

    if ENABLE_IV:
        step += 1
        print(f"\n[{step}/{total_steps}] Invoice / IV (รายได้)...")
        iv_list, iv_lines = fetch_invoice_module()
        datasets["Invoice_Summary"] = iv_list
        datasets["Invoice_Items"]   = iv_lines

    has_data = any(len(v) > 0 for v in datasets.values())
    if has_data:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        prefix = "TRCLOUD_PO_Export" if ENABLE_PO and not any(
            [ENABLE_AP, ENABLE_PV, ENABLE_XEXPENSE, ENABLE_IV]
        ) else "TRCLOUD_Master_Export"
        master_file = export_master_excel(datasets, ts, filename_prefix=prefix)

        if ENABLE_DASHBOARD and ENABLE_AP and ENABLE_PV:
            dashboard_file = os.path.join(OUTPUT_DIR, f"Dashboard_Data_Model_{ts}.xlsx")
            generate_dashboard_model(master_file, dashboard_file)
            print(f"   Dashboard : {os.path.basename(dashboard_file)}")

        print("\n🎉 เสร็จสมบูรณ์!")
        print(f"   Raw Data  : {os.path.basename(master_file)}")
        if ENABLE_PO and po_list:
            export_po_json(po_list, po_lines)
    else:
        print("⚠️ ไม่พบข้อมูลในช่วงเวลาที่ระบุ")